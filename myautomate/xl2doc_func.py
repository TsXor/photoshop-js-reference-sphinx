import openpyxl, sys

template = '''
.. function:: <FUNC>
	
	<DESC>
	
	:Parameters:(MULTI)
		:<PARA_NAME>: :class:`<PARA_TYPE>`(MULTI)
	
	:Returns: :class:`<RETURN TYPE>`
'''

def insert_temp(func, desc, paras, rtype):
    t = template
    t = t.replace('<FUNC>', func)
    t = t.replace('<DESC>', desc)
    t = t.replace(':class:`<RETURN TYPE>`', '`undefined`') \
        if rtype is None \
        else t.replace('<RETURN TYPE>', rtype)
    tm = t.split('(MULTI)')
    mul = tm[1]
    mullist = []
    if paras is None:
        mullist = [' `null`']
    else:
        for pname, ptype in paras:
            cmul = mul
            cmul = cmul.replace('<PARA_NAME>', pname)
            cmul = cmul.replace('<PARA_TYPE>', ptype)
            mullist.append(cmul)
    tm = [tm[0], *mullist, tm[2]]
    t = ''.join(tm)
    return t

def loadws(ws):
    nrow = ws.max_row
    ncolumn = ws.max_column
    return [['' if (v := ws.cell(row=i, column=x).value) is None else v for x in range(1, ncolumn+1)] for i in range(1, nrow+1)]

def main():
    wb = openpyxl.load_workbook(sys.argv[1])
    ws = wb['parsed']
    dlist = loadws(ws)
    for l in dlist:
        v1 = l[1-1]
        v2 = v2.split('\n') if (v2 := l[2-1]) else None
        v3 = v3.split('\n') if (v3 := l[3-1]) else None
        v4 = v4 if (v4 := l[4-1]) else None
        v5 = l[5-1]
        paras = [(v2[i], v3[i]) for i in range(len(v2))] if v2 else None
        l = [v1, v5, paras, v4]
        doc = insert_temp(*l)
        print(doc)



if __name__ == '__main__':
    main()