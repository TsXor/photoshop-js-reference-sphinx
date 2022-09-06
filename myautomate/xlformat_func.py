import openpyxl, sys, re
from copy import *

js_type_keywords = ('number', 'string', 'boolean')
array_kwd_matcher = re.compile(r'array of (\d* )?\w*')
or_kwd_matcher = re.compile(r'\w* or \w*')

def extract(ws, decider):
    decider_range = []
    for rect in ws.merged_cells.ranges:
        t, b, l, r = rect.min_row, rect.max_row, rect.min_col, rect.max_col
        if l == decider and r == decider:
            decider_range.append((t-1, b))
    decider_range.sort(key=lambda x: x[0])
    nrow = ws.max_row
    ncolumn = ws.max_column
    decider_n = [i for r in decider_range for i in range(*r)]
    range_patches = []
    for i in range(nrow):
        if not i in decider_n:
            patch = (i, i+1)
            range_patches.append(patch)
    decider_range.extend(range_patches)
    decider_range.sort(key=lambda x: x[0])
    vessel = []
    for i in range(1, ncolumn+1):
        coldump = []
        if i == decider:
            coldump = [ws.cell(column=i, row=start+1).value for start, end in decider_range]
        else:
            coldump = [cell if (cell := ws.cell(column=i, row=j).value) else '' for j in range(1, nrow+1)]
            coldump = [[coldump[n] for n in range(*drange)] for drange in decider_range]
            coldump = [' '.join([s for s in lcell if s]) for lcell in coldump]
        vessel.append(coldump)
    return vessel

def delchar(string, chars):
    for char in chars:
        string = string.replace(char, '')
    return string

def f2argkeys(x):
    target = x[0]
    funcs = [delchar(s, ' ').replace(',', ', ').replace('][', '') for s in target]
    target = [delchar(s, '[] ') for s in funcs]
    bracket_matcher = re.compile(r'\(.*\)')
    try:
        return (funcs, ['\n'.join(bracket_matcher.search(s).group()[1:-1].split(',')) if s else s for s in target])
    except:
        print('Exception in parsing:')
        for s in target:
            m = bracket_matcher.search(s)
            if m is None:
                print('Unable to parse: ', s)
        raise ValueError

def regexpick(string, matcher, replace):
    s = string
    kwds = []
    while True:
        search = matcher.search(s)
        if search is None:
            break
        kwd = search.group()
        kwds.append(kwd)
        s = s.replace(kwd, replace, 1)
    return (s, kwds)

def putback(alist, kwds, replace):
    alistc = copy(alist)
    kwdsc = copy(kwds)
    for i in range(len(alistc)):
        e = alistc[i]
        if e == replace:
            alistc[i] = kwdsc[0]
            kwdsc.pop(0)
    return alistc

def splitargtypes(string):
    s = string
    s, arrkwds = regexpick(s, array_kwd_matcher, 'ARR')
    s, orkwds = regexpick(s, or_kwd_matcher, 'LOR')
    split = [l for l in s.split(' ') if l]
    split = putback(split, orkwds, 'LOR')
    split = putback(split, arrkwds, 'ARR')
    return split

def r2argtypes(x):
    target = x[1]
    return ['\n'.join(splitargtypes(s)) if s else s for s in target]

def dump2ws(ws, x):
    for i in range(len(x[0])):
        for j in range(len(x)):
            tchr = 'ABCDEFGHIJKLMNOPQRSTUVWXYZ'[j]
            ws['%s%d'%(tchr, i+1)] = x[j][i]
            ws['%s%d'%(tchr, i+1)].alignment = openpyxl.styles.Alignment(wrapText=True)

def main():
    wb = openpyxl.load_workbook(sys.argv[1])
    ws = wb.active
    x = extract(ws, 3)
    funcs, argkeys = f2argkeys(x)
    argtypes = r2argtypes(x)
    x = [funcs, argkeys, argtypes, x[2], x[3]]
    wsr = wb.create_sheet('sheet')
    wsr.title = 'parsed'
    dump2ws(wsr, x)
    wb.save(sys.argv[1])



if __name__ == '__main__':
    main()